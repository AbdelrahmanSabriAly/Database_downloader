import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import pickle
import os
import cv2
import streamlit as st
import gdown
from tqdm import tqdm
from openpyxl import load_workbook


def create_keyfile_dict():
    variables_keys = {
        "type": st.secrets["TYPE"],
        "project_id": st.secrets["PROJECT_ID"],
        "private_key_id": st.secrets["PRIVATE_KEY_ID"],
        "private_key": st.secrets["PRIVATE_KEY"],
        "client_email": st.secrets["CLIENT_EMAIL"],
        "client_id": st.secrets["CLIENT_ID"],
        "auth_uri": st.secrets["AUTH_URI"],
        "token_uri": st.secrets["TOKEN_URI"],
        "auth_provider_x509_cert_url": st.secrets["AUTH_PROVIDER_X509_CERT_URL"],
        "client_x509_cert_url": st.secrets["CLIENT_X509_CERT_URL"],
        "universe_domain" : st.secrets["UNIVERSE_DOMAIN"]
    }
    return variables_keys



face_dict = {}
mac_dict = {}
dirname = ""

# Google Sheet details
SHEET_NAME = 'Your Sheet Name'
ID_COLUMN_NAME = 'ID'
TIMESTAMP_COLUMN_NAME = 'Timestamp'


def recognize_face(image, face_detector, face_recognizer, file_name=None):
    channels = 1 if len(image.shape) == 2 else image.shape[2]
    if channels == 1:
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
    if channels == 4:
        image = cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)

    if image.shape[0] > 1000:
        image = cv2.resize(image, (0, 0),
                           fx=500 / image.shape[0], fy=500 / image.shape[0])

    height, width, _ = image.shape
    face_detector.setInputSize((width, height))
    try:
        _, faces = face_detector.detect(image)
        if file_name is not None:
            assert len(faces) > 0, f'the file {file_name} has no face'

        faces = faces if faces is not None else []
        features = []
        for face in faces:

            aligned_face = face_recognizer.alignCrop(image, face)
            feat = face_recognizer.feature(aligned_face)

            features.append(feat)
        return features, faces
    except Exception as e:
        print(e)
        print(file_name)
        return None, None


def load_models():
    # Init models face detection & recognition
    weights = "required_files/face_detection_yunet_2023mar_int8.onnx"
    face_detector = cv2.FaceDetectorYN_create(weights, "", (0, 0))
    face_detector.setScoreThreshold(0.87)

    weights = "required_files/face_recognition_sface_2021dec_int8.onnx"
    face_recognizer = cv2.FaceRecognizerSF_create(weights, "")
    return face_detector,face_recognizer

def download_image_from_drive(url, output_dir):
    output_path = os.path.join(output_dir, '')
    gdown.download(url, output_path, quiet=False)
    output =os.path.join(output_dir, os.listdir(output_dir)[0])
    return output

def get_pickled_excel():
    # Load the Excel file
    excel_file_path = "utils\\2025.xlsx"
    workbook = load_workbook(excel_file_path)

    # Assuming you want to work with the first sheet
    sheet = workbook.active

    # Process your Excel data (read/write as needed)
    data_to_pickle = []

    for row in sheet.iter_rows(values_only=True):
        data_to_pickle.append(row)

    # Close the workbook
    workbook.close()

    return data_to_pickle


def load_forms_responses(face_detector, face_recognizer,temp,year):

    file_name = f"{year}_data.pkl"
    if os.path.exists(file_name):
        os.remove(file_name)

    eece_2025 = get_pickled_excel() 

    student_data = []

    sheet_name = "Attendance monitoring (Responses)"

    # Authorize with Google Sheets API using credentials
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_dict(create_keyfile_dict(), scope)
    client = gspread.authorize(creds)
    # Open the Google Sheet by its name
    sheet = client.open(sheet_name).sheet1
    sheet = find_and_remove_duplicates(sheet)
    data = sheet.get_all_values()  # Get all the values from the sheet

    num_rows = len(data)  # Get the number of rows in the sheet

    # Get all the values from the sheet
    data = sheet.get_all_records()

    my_bar = st.progress(0.0, text="Processing ...")
    added_value = 1.0/num_rows
    counter = 0
    output_dir = 'IMAGES'
    # Process the data
    for row in tqdm(data):
        timestamp = row['Timestamp']
        mac = row['MAC Addresses']
        id = row['ID']
        name_in_arabic = row['Name in Arabic']
        image_url = row['Image']
        image_url = image_url.replace("open","uc")

        output = download_image_from_drive(image_url, output_dir)
        image = cv2.imread(output)
        # Process the image using face recognition functions
        feats, faces = recognize_face(image, face_detector, face_recognizer)

        if faces is None:
            continue

        # Extract user_id from the uploaded file's name
        face_dict[id] = feats[0]
        mac_dict[id] = mac
        student_data.append({"Name": name_in_arabic, "ID": id})

        os.remove(output)
        counter+=added_value
        my_bar.progress(counter, text="Processing ...")
        
    st.success(f'There are {len(face_dict)} students')
    my_bar.progress(1.0, text="Done")
        
        
    temp.empty()
    df = pd.DataFrame(student_data)
    
    pickled_list = [face_dict,mac_dict,eece_2025]
    with open(file_name, "wb") as file:  # Use 'wb' mode for writing binary data
        pickle.dump(pickled_list, file)


    st.download_button(
            label="Click here to download",
            data=open(file_name, "rb").read(),
            file_name=file_name,
            mime="application/octet-stream",
        )

    df = pd.DataFrame(student_data)
    st.table(df)



def find_and_remove_duplicates(sheet):
    all_records = sheet.get_all_records()
    ids_to_records = {}
    rows_to_delete = []

    # Find the latest entry for each ID
    for idx, record in enumerate(all_records):
        current_id = record[ID_COLUMN_NAME]
        if current_id in ids_to_records:
            existing_record = ids_to_records[current_id]
            if existing_record[TIMESTAMP_COLUMN_NAME] < record[TIMESTAMP_COLUMN_NAME]:
                rows_to_delete.append(existing_record['row'])
                ids_to_records[current_id] = {'row': idx + 2, TIMESTAMP_COLUMN_NAME: record[TIMESTAMP_COLUMN_NAME]}
            else:
                rows_to_delete.append(idx + 2)
        else:
            ids_to_records[current_id] = {'row': idx + 2, TIMESTAMP_COLUMN_NAME: record[TIMESTAMP_COLUMN_NAME]}

    # Delete rows with older timestamps
    if rows_to_delete:
        for row_num in reversed(rows_to_delete):
            sheet.delete_row(row_num)

    return sheet
